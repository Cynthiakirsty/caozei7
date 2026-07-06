"""双表经营报告导出。"""
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, PatternFill

from modules.ai_advisor import generate_advice
from modules.basic_analysis import analysis_to_rows, generate_detailed_analysis
from modules.churn import score_churn
from modules.metrics import calculate_metrics
from modules.segmentation import segment_users


def build_excel_report(full_df: pd.DataFrame, user_df: pd.DataFrame) -> bytes:
    labels = {
        "total_deposit":"总充值", "total_withdraw":"总提现", "surplus":"充减提",
        "surplus_rate":"盈余率", "total_bet":"有效投注", "rtp":"RTP",
        "kill_rate":"杀率", "activity_cost":"彩金金额",
        "activity_cost_ratio":"彩金占比", "arppu":"ARPPU", "new_arpu":"新增ARPU",
        "new_arppu":"新增ARPPU", "first_recharge_rate":"首充复充率",
        "old_arppu":"老用户ARPPU", "first_deposit_surplus_rate":"首充盈余率",
        "new_users":"新增用户数", "retention_d1":"次日留存",
        "retention_d3":"3日留存", "retention_d7":"7日留存",
        "retention_d15":"15日留存", "retention_d30":"30日留存",
    }
    metrics = calculate_metrics(full_df)
    summary = pd.DataFrame([{"指标": labels[k], "数值": v} for k, v in metrics.items() if k in labels])
    advice = pd.DataFrame(generate_advice(full_df)).rename(
        columns={"level":"优先级", "title":"建议主题", "detail":"建议内容"}
    )
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="指标摘要")
        full_df.to_excel(writer, index=False, sheet_name="全盘数据")
        user_df.to_excel(writer, index=False, sheet_name="用户明细")
        segment_users(user_df).to_excel(writer, index=False, sheet_name="用户分层")
        score_churn(user_df).head(1000).to_excel(writer, index=False, sheet_name="流失预警")
        analysis_to_rows(generate_detailed_analysis(full_df, user_df)).to_excel(
            writer, index=False, sheet_name="基础决策建议"
        )
        advice.to_excel(writer, index=False, sheet_name="运营建议")
        for ws in writer.book.worksheets:
            ws.freeze_panes, ws.auto_filter.ref = "A2", ws.dimensions
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2563EB")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    36, max(12, max(len(str(c.value or "")) for c in col)+2)
                )
    return output.getvalue()
