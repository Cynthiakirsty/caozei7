"""双表 Excel 数据加载、校验、示例数据及 SQLite 持久化。"""
from pathlib import Path
import sqlite3

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill

FULL_COLUMNS_CN = [
    "日期", "新增用户数", "日活跃玩家数", "老玩家日活", "充值人数", "充值金额",
    "提现金额", "充减提", "盈余率", "有效投注", "首充人数", "首充金额",
    "首充盈余率", "首充提现占比", "首充付费率", "首充复充率", "新增ARPU",
    "新增ARPPU", "老用户充值人数", "老用户充值金额", "老用户付费率",
    "老用户ARPU", "老用户ARPPU", "老用户提现金额", "老用户提现占比",
    "老用户充减提", "老用户盈余率", "付费率", "ARPPU", "ARPU",
    "次日留存", "3日留存", "7日留存", "15日留存", "30日留存",
    "rtp", "杀率", "彩金金额", "彩金占比",
]
USER_COLUMNS_CN = [
    "用户UID", "VIP等级", "渠道", "注册时间", "最后登录时间", "首充时间",
    "累计流水", "累计充值", "累计提现", "Cash余额", "JCoin余额",
]

FULL_MAP = {
    "日期":"date", "新增用户数":"new_users", "日活跃玩家数":"dau",
    "老玩家日活":"old_dau", "充值人数":"paying_users", "充值金额":"deposit",
    "提现金额":"withdraw", "充减提":"surplus", "盈余率":"surplus_rate",
    "有效投注":"bet", "首充人数":"first_deposit_users", "首充金额":"first_deposit",
    "首充盈余率":"first_deposit_surplus_rate", "首充提现占比":"first_withdraw_ratio",
    "首充付费率":"first_pay_rate", "首充复充率":"first_recharge_rate",
    "新增arpu":"new_arpu", "新增arppu":"new_arppu",
    "老用户充值人数":"old_paying_users", "老用户充值金额":"old_deposit",
    "老用户付费率":"old_pay_rate", "老用户arpu":"old_arpu",
    "老用户arppu":"old_arppu", "老用户提现金额":"old_withdraw",
    "老用户提现占比":"old_withdraw_ratio", "老用户充减提":"old_surplus",
    "老用户盈余率":"old_surplus_rate", "付费率":"pay_rate", "arppu":"arppu",
    "arpu":"arpu", "次日留存":"retention_d1", "3日留存":"retention_d3",
    "7日留存":"retention_d7", "15日留存":"retention_d15",
    "30日留存":"retention_d30", "rtp":"rtp", "杀率":"kill_rate",
    "彩金金额":"bonus", "彩金占比":"bonus_ratio",
}
USER_MAP = {
    "用户uid":"user_id", "vip等级":"vip_level", "渠道":"channel",
    "注册时间":"register_date", "最后登录时间":"last_login_date",
    "首充时间":"first_deposit_date", "累计流水":"total_bet",
    "累计充值":"total_deposit", "累计提现":"total_withdraw",
    "cash余额":"cash_balance", "jcoin余额":"jcoin_balance",
}
FULL_COLUMNS = list(FULL_MAP.values())
USER_COLUMNS = list(USER_MAP.values())
FULL_RATE_COLUMNS = [c for c in FULL_COLUMNS if c.endswith(("rate", "ratio"))] + [
    "retention_d1", "retention_d3", "retention_d7", "retention_d15",
    "retention_d30", "rtp",
]
USER_DATE_COLUMNS = ["register_date", "last_login_date", "first_deposit_date"]


def _normalize_headers(columns, mapping):
    """中文字段大小写不敏感，同时允许标准英文字段。"""
    return [mapping.get(str(c).strip().lower(), str(c).strip().lower()) for c in columns]


def generate_full_sample(days: int = 90, seed: int = 42) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    dates = pd.date_range(end=pd.Timestamp.today().normalize(), periods=days)
    rows = []
    for date in dates:
        new_users = int(rng.integers(80, 260))
        dau = int(rng.integers(900, 2100))
        old_dau = max(0, dau-new_users)
        paying = int(dau*rng.uniform(.18, .32))
        deposit = round(paying*rng.uniform(180, 360), 2)
        withdraw = round(deposit*rng.uniform(.52, .83), 2)
        surplus = deposit-withdraw
        bet = round(deposit*rng.uniform(3.5, 7.5), 2)
        first_users = int(new_users*rng.uniform(.15, .28))
        first_deposit = round(first_users*rng.uniform(120, 260), 2)
        old_paying = max(0, paying-first_users)
        old_deposit = max(0, deposit-first_deposit)
        old_withdraw = round(withdraw*rng.uniform(.62, .85), 2)
        bonus = round(deposit*rng.uniform(.04, .13), 2)
        rtp = float(rng.uniform(.88, .97))
        rows.append([
            date, new_users, dau, old_dau, paying, deposit, withdraw, surplus,
            surplus/deposit, bet, first_users, first_deposit, rng.uniform(.12, .35),
            rng.uniform(.35, .72), first_users/new_users, rng.uniform(.18, .42),
            deposit/new_users, first_deposit/first_users if first_users else 0,
            old_paying, old_deposit, old_paying/old_dau if old_dau else 0,
            old_deposit/old_dau if old_dau else 0,
            old_deposit/old_paying if old_paying else 0, old_withdraw,
            old_withdraw/old_deposit if old_deposit else 0, old_deposit-old_withdraw,
            (old_deposit-old_withdraw)/old_deposit if old_deposit else 0,
            paying/dau, deposit/paying if paying else 0, deposit/dau,
            rng.uniform(.30, .48), rng.uniform(.25, .42), rng.uniform(.20, .36),
            rng.uniform(.15, .30), rng.uniform(.10, .24), rtp,
            (bet-bet*rtp)/deposit, bonus, bonus/deposit,
        ])
    return pd.DataFrame(rows, columns=FULL_COLUMNS)


def generate_user_sample(users: int = 500, seed: int = 43) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    today = pd.Timestamp.today().normalize()
    rows = []
    for i in range(1, users+1):
        register = today-pd.Timedelta(days=int(rng.integers(1, 240)))
        last_login = today-pd.Timedelta(days=int(rng.integers(0, 45)))
        paid = rng.random() < .72
        first_deposit = register+pd.Timedelta(days=int(rng.integers(0, 8))) if paid else pd.NaT
        deposit = round(float(rng.gamma(2.2, 1200)), 2) if paid else 0
        bet = round(deposit*rng.uniform(2, 12), 2)
        withdraw = round(deposit*rng.uniform(.2, .95), 2)
        vip = f"VIP{min(8, int(deposit//3000))}"
        rows.append([
            f"UID{i:06d}", vip, rng.choice(["自然量", "代理A", "代理B", "广告投放"]),
            register, last_login, first_deposit, bet, deposit, withdraw,
            round(max(0, deposit-withdraw)*rng.uniform(.05, .35), 2),
            round(rng.uniform(0, 800), 2),
        ])
    return pd.DataFrame(rows, columns=USER_COLUMNS)


def _write_styled_excel(df: pd.DataFrame, path: Path, chinese_headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    output = df.copy()
    output.columns = chinese_headers
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        output.to_excel(writer, index=False, sheet_name="数据")
        ws = writer.book["数据"]
        ws.freeze_panes, ws.auto_filter.ref = "A2", ws.dimensions
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                20, max(12, max(len(str(c.value or "")) for c in col)+2)
            )


def ensure_sample_files(data_dir: Path = Path("data")) -> tuple[Path, Path]:
    full_path, user_path = data_dir/"sample_full_data.xlsx", data_dir/"sample_user_data.xlsx"
    _write_styled_excel(generate_full_sample(), full_path, FULL_COLUMNS_CN)
    _write_styled_excel(generate_user_sample(), user_path, USER_COLUMNS_CN)
    return full_path, user_path


def validate_full_data(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    df = df.copy()
    df.columns = _normalize_headers(df.columns, FULL_MAP)
    missing = [c for c in FULL_COLUMNS if c not in df.columns]
    if missing:
        reverse = {v:k for k,v in FULL_MAP.items()}
        raise ValueError("全盘数据缺少字段："+"、".join(reverse.get(c, c) for c in missing))
    df = df[FULL_COLUMNS]
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    for col in FULL_COLUMNS[1:]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    bad = df["date"].isna()
    warnings = [f"已移除 {int(bad.sum())} 条无效日期记录"] if bad.any() else []
    return df.loc[~bad].sort_values("date").reset_index(drop=True), warnings


def validate_user_data(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    df = df.copy()
    df.columns = _normalize_headers(df.columns, USER_MAP)
    missing = [c for c in USER_COLUMNS if c not in df.columns]
    if missing:
        reverse = {v:k for k,v in USER_MAP.items()}
        raise ValueError("用户明细缺少字段："+"、".join(reverse.get(c, c) for c in missing))
    df = df[USER_COLUMNS]
    df["user_id"] = df["user_id"].astype(str).str.strip()
    for col in USER_DATE_COLUMNS:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    for col in ["total_bet", "total_deposit", "total_withdraw", "cash_balance", "jcoin_balance"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).clip(lower=0)
    bad = df["user_id"].isin(["", "nan", "None"])
    warnings = [f"已移除 {int(bad.sum())} 条无 UID 记录"] if bad.any() else []
    return df.loc[~bad].drop_duplicates("user_id", keep="last").reset_index(drop=True), warnings


def load_full_excel(source) -> pd.DataFrame:
    return validate_full_data(pd.read_excel(source, engine="openpyxl"))[0]


def load_user_excel(source) -> pd.DataFrame:
    return validate_user_data(pd.read_excel(source, engine="openpyxl"))[0]


def save_to_sqlite(full_df: pd.DataFrame, user_df: pd.DataFrame,
                   db_path: Path = Path("data/operations.db")) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        full_df.to_sql("full_operation_data", conn, if_exists="replace", index=False)
        user_df.to_sql("user_detail_data", conn, if_exists="replace", index=False)
